"""Automatyczna synchronizacja arkusza 'Przeglądy' z Supabase.

Wymagane zmienne środowiskowe:
  SUPABASE_URL=https://twoj-projekt.supabase.co
  SUPABASE_SERVICE_ROLE_KEY=...  (sekretny klucz, nigdy do aplikacji WWW)
"""
import datetime as dt
import hashlib
import json
import os
import sys
import unicodedata
import urllib.error
import urllib.request
from pathlib import Path

import openpyxl


def norm(value):
    text = unicodedata.normalize("NFKD", str(value or ""))
    return "".join(c for c in text if not unicodedata.combining(c)).lower().strip()


def date(value):
    if not value:
        return None
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%Y-%m-%d")
    try:
        return dt.datetime.fromisoformat(str(value)).strftime("%Y-%m-%d")
    except ValueError:
        return None


def source_id(city, local, inspection_type):
    """Stały identyfikator rekordu z Excela, niezależny od numeru wiersza."""
    identity = "|".join(norm(value) for value in (city, local, inspection_type))
    return "excel-" + hashlib.sha256(identity.encode("utf-8")).hexdigest()[:24]


def main():
    url = os.getenv("SUPABASE_URL", "").rstrip("/")
    key = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")
    excel = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).with_name("przeglądy.xlsx")
    if not url or not key:
        raise SystemExit("Brak SUPABASE_URL lub SUPABASE_SERVICE_ROLE_KEY w zmiennych środowiskowych.")
    if not excel.exists():
        raise SystemExit(f"Nie znaleziono pliku Excel: {excel}")

    book = openpyxl.load_workbook(excel, data_only=True, read_only=True)
    sheet = next((book[n] for n in book.sheetnames if "przeglad" in norm(n)), None)
    if sheet is None:
        raise SystemExit("Nie znaleziono arkusza 'Przeglądy'.")
    headers = [norm(c.value) for c in next(sheet.iter_rows(min_row=1, max_row=1))]

    def value(row, phrase):
        index = next((i for i, h in enumerate(headers) if phrase in h), None)
        return row[index] if index is not None else None

    records = []
    for excel_row, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        city, local, kind = value(cells, "miasto"), value(cells, "nr lokalu"), value(cells, "rodzaj przegl")
        if not any((city, local, kind)):
            continue
        months = value(cells, "wazny przez") or 12
        try:
            months = int(months)
        except (TypeError, ValueError):
            months = 12
        city, local, kind = str(city or "").strip(), str(local or "").strip(), str(kind or "").strip()
        if not (city and local and kind):
            print(f"Pomijam wiersz {excel_row}: brakuje miasta, lokalu lub rodzaju przeglądu.")
            continue
        records.append({
            "id": source_id(city, local, kind), "source_id": source_id(city, local, kind),
            "city": city, "local": local, "type": kind,
            "done": date(value(cells, "data wykon")), "months": months,
            "protocol_date": date(value(cells, "data dodania protokol")),
            "notes": str(value(cells, "uwagi") or "").strip(),
        })
    list_sheet = next((book[n] for n in book.sheetnames if norm(n) == "listy"), None)
    type_names = []
    if list_sheet:
        list_headers = [norm(c.value) for c in next(list_sheet.iter_rows(min_row=1, max_row=1))]
        type_index = next((i for i, header in enumerate(list_headers) if "rodzaj przegl" in header), None)
        if type_index is not None:
            for row in list_sheet.iter_rows(min_row=2, values_only=True):
                name = str(row[type_index] or "").strip()
                if name and name not in type_names:
                    type_names.append(name)
    if not type_names:
        type_names = list(dict.fromkeys(record["type"] for record in records if record["type"]))

    locations_sheet = next((book[n] for n in book.sheetnames if norm(n) == "lokale"), None)
    locations = []
    if locations_sheet:
        location_headers = [norm(c.value) for c in next(locations_sheet.iter_rows(min_row=1, max_row=1))]
        city_index = next((i for i, header in enumerate(location_headers) if header == "miasto"), None)
        local_index = next((i for i, header in enumerate(location_headers) if "nr lokalu" in header), None)
        if city_index is not None and local_index is not None:
            seen_locations = set()
            for row in locations_sheet.iter_rows(min_row=2, values_only=True):
                city = str(row[city_index] or "").strip()
                local = str(row[local_index] or "").strip()
                if city and local and (city, local) not in seen_locations:
                    locations.append({"city": city, "local": local})
                    seen_locations.add((city, local))
    if not locations:
        locations = list({(record["city"], record["local"]) for record in records if record["city"] and record["local"]})
        locations = [{"city": city, "local": local} for city, local in locations]

    def upsert(table, rows, conflict):
        if not rows:
            return
        request = urllib.request.Request(
            f"{url}/rest/v1/{table}?on_conflict={conflict}", data=json.dumps(rows).encode(), method="POST",
            headers={"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json", "Prefer": "resolution=merge-duplicates,return=minimal"},
        )
        try:
            with urllib.request.urlopen(request) as response:
                return response.status
        except urllib.error.HTTPError as error:
            raise SystemExit(f"Supabase zwrócił HTTP {error.code}: {error.read().decode(errors='replace')}")

    # Pierwsze uruchomienie po aktualizacji zachowuje dotychczasowe identyfikatory
    # rekordów (seed-...) i dopisuje im trwały source_id.
    existing_request = urllib.request.Request(
        f"{url}/rest/v1/inspections?select=id,city,local,type,source_id",
        headers={"apikey": key, "Authorization": f"Bearer {key}"},
    )
    try:
        with urllib.request.urlopen(existing_request) as response:
            existing = json.load(response)
    except urllib.error.HTTPError as error:
        raise SystemExit(f"Nie można odczytać istniejących przeglądów (HTTP {error.code}): {error.read().decode(errors='replace')}")
    legacy_ids = {
        source_id(row.get("city"), row.get("local"), row.get("type")): row["id"]
        for row in existing if not row.get("source_id")
    }
    for record in records:
        record["id"] = legacy_ids.get(record["source_id"], record["id"])

    inspections_status = upsert("inspections", records, "source_id")
    types_status = upsert("inspection_types", [{"name": name} for name in type_names], "name")
    locations_status = upsert("locations", locations, "city,local")
    print(f"Zsynchronizowano {len(records)} wpisów (HTTP {inspections_status}), {len(type_names)} rodzajów przeglądów (HTTP {types_status}) i {len(locations)} lokali (HTTP {locations_status}).")


if __name__ == "__main__":
    main()
